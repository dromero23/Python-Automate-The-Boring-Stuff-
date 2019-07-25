#!python3
#xltoCSV.py - converts the excel files into csv for each sheet with the name:[workbookname]_[sheetname].csv
#input: Excel file 
#output: CSV file(s)

import openpyxl, csv, os,logging 
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s- %(message)s')
logging.disable(logging.CRITICAL)

def main():
	for filename in os.listdir('.'):
		if not filename.endswith('.xlsx'):
			continue
		logging.debug(filename)
		wb = openpyxl.load_workbook(filename)
		fn = filename.split('.')[0] #get the filename
		for sheetname in wb.sheetnames:
			sheet = wb.get_sheet_by_name(sheetname)
			outputFile = open(fn+'_'+sheetname+'.csv', 'w', newline ='')
			outputWriter = csv.writer(outputFile)
			for rowNum in range(1,sheet.max_row+1):
				rowData =[]
				for colNum in range(1,sheet.max_column+1):
					rowData.append(str(sheet[get_column_letter(colNum)+str(rowNum)].value))
				outputWriter.writerow(rowData)
			outputFile.close()
		
if __name__ == '__main__':
	main()