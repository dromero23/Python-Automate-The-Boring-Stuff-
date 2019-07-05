#!python3
#SStoTXT.py - opens a spreadsheet and writes the cells of column A into one text file, the cells of column B into another text file, etc.
#input: excel file
#output: N text files, in which N represents the number of columns from the input excel spreadsheet

import openpyxl, sys 
from openpyxl.utils import get_column_letter


def main():
	#Read spreasheet
	if len(sys.argv)<2:
		print('Please enter a spreadsheet filename from the local folder')
		sys.exit()
	wb = openpyxl.load_workbook(sys.argv[1])
	fileName = sys.argv[1].split('.')[0] #get the filename
	sheet = wb.active
	#TODO: Seperate columns and read all the column's rows and export to a text file until all columns have been read and save
	for colNum in range(1,sheet.max_column+1):
		newFile = open(fileName+'_text'+str(colNum)+'.txt','w')
		for rowNum in range(1,sheet.max_row+1):
			newFile.write(str(sheet[get_column_letter(colNum)+str(rowNum)].value)+'\n')
		newFile.close()
	print('Done!')

if __name__ =='__main__':
	main()