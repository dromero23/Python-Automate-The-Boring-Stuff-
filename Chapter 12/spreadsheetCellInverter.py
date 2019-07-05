#!python3
#spreadsheetCellInverter.py - inverts the column and row of the cells in a spreadsheet
#input: spreadsheet filename
#ouput: Creates an inverted input spreadsheet

import openpyxl, sys 
from openpyxl.utils import get_column_letter, column_index_from_string

def main():
	if len(sys.argv)<2:
		print('Correct usage: py spreadsheetCellInverter.py filename')
		sys.exit()
	wb = openpyxl.load_workbook(sys.argv[1])
	nwb = openpyxl.Workbook()
	sheet = wb.active
	nsheet = nwb.active
	for rowNum in range(1,sheet.max_row+1):
		for colNum in range(1,sheet.max_column+1):
			nsheet[get_column_letter(rowNum)+str(colNum)]=sheet[get_column_letter(colNum)+str(rowNum)].value
	nwb.save('inverted_'+sys.argv[1])

if __name__ == '__main__':
	main()