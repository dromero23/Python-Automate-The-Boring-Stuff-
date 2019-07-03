#!python3
#multiplicationTable.py - takes a number N from the command line and creates an N×N multiplication table in an Excel spreadsheet: ex. py multiplicationTable.py n, in which n represents an unsigned integer
 
import openpyxl, sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
def main():
	if len(sys.argv) <2:
		print('Usage: python stripfunc.py [integer]')
		sys.exit()
	wb = openpyxl.Workbook()
	sheet = wb.active
	try: 
		n = int(sys.argv[1])
	except ValueError:
		print('You\'ve enter a non-integer input')
		print('Usage: python stripfunc.py [integer]')
		sys.exit()
	fontObj = Font(bold = True)
	for column in range(1,n+1):
		sheet[get_column_letter(column+1)+'1'].font = fontObj
		sheet[get_column_letter(column+1)+'1'] = column
		sheet['A'+str(column+1)].font = fontObj 
		sheet['A'+str(column+1)]=column
		for row in range(1,n+1):
			sheet[get_column_letter(column+1)+str(row+1)]= '='+get_column_letter(column+1)+'1*'+'A'+str(row+1)
	wb.save('multiplicationTable.xlsx')

			
if __name__=='__main__':
	main()