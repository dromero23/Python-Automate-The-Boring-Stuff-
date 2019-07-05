#!python3 
#TXTtoSS.py - opens a spreadsheet and writes the context of your text files into a spreadsheet, the columns represents the number of textFiles being read and the rows is the context with newline delimited
#input - text files
#output - spreadsheet

import openpyxl
from openpyxl.utils import get_column_letter

def main():
	fList = []
	wb = openpyxl.Workbook()
	sheet = wb.active
	while(True):
		user = input('Enter your filename or press enter to quit: ')
		if (user ==''):
			break
		fList.append(user)
	#the for loop represents the column
	for textFile in range(len(fList)): 
		try:
			newFile = open(fList[textFile],'r')
		except FileNotFoundError: 
			print('"'+fList[textFile] +'" does not exist')
			continue
		rowList = newFile.readlines()
		#for the row
		for rowNum in range(len(rowList)):
			sheet[get_column_letter(textFile+1)+str(rowNum+1)]=rowList[rowNum]
	wb.save('ss.xlsx')
		
if __name__=='__main__':
	main()