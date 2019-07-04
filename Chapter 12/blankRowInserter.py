#!python3
#blankRowInserter.py - takes a spreadsheet and inserts M blank rows from the Nth row.
#Usage: blankRowInserter.py [integer - n] [integer - m] fileName
#output: fileName will be edited, for testing purpose for now it will be called newFile.xlsx

'''
You can write this program by reading in the contents of the spreadsheet. Then, when writing out the new spreadsheet, use a for loop to copy the first N lines. For the remaining lines, add M to the row number in the output spreadsheet.
'''

import openpyxl, sys
from openpyxl.utils import get_column_letter

def main():
	if len(sys.argv) <3:
		print('Correct usage: blankRowInserter.py [integer] [integer] fileName')
		sys.exit()
	wb = openpyxl.load_workbook(sys.argv[3])
	nwb = openpyxl.Workbook()
	sheet = wb.active
	nsheet = nwb.active
	try: 
		n = int(sys.argv[1])
		m = int(sys.argv[2])
	except ValueError:
		print('You\'ve enter a non-integer input')
		print('Usage: blankRowInserter.py [integer] [integer] fileName')
	
	#READ the data 
	#insert all values up to the nth row
	for rowNum in range(1,n):
		for colNum in range(1,sheet.max_column+1):
			 nsheet[get_column_letter(colNum)+str(rowNum)] = sheet[get_column_letter(colNum)+str(rowNum)].value
	#TODO: insert the remaining data from nth+m  up to the last row
	for rowNum in range(n,sheet.max_row+1):
		for colNum in range(1,sheet.max_column+1):
			nsheet[get_column_letter(colNum)+str(rowNum+m)] = sheet[get_column_letter(colNum)+str(rowNum)].value
	nwb.save('new_'+sys.argv[3])

if __name__=='__main__':
	main()